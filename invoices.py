import os
from openpyxl import Workbook
import pdfplumber 
import re
from datetime import datetime

print ('Iniciando execução')

directory = 'in-voices'
files = os.listdir(directory)
files_qtd = len(files)

if files_qtd == 0:
    raise Exception('Sem arquivos no diretório.')

wb = Workbook()
ws = wb.active
ws.title = 'Invoice Imports'

ws['A1'] = 'Invoice #'
ws['B1'] = 'Date'
ws['C1'] = 'File Name'
ws['D1'] = 'Status'

print('entrei no loop')

last_empty_line = 1

while ws['A1' + str(last_empty_line)].value is not None:
    last_empty_line += 1

for file in files:
    with pdfplumber.open(directory + "/" + file) as pdf:
        first_page = pdf.pages[0]
        pdf_text = first_page.extract_text()
        print(pdf_text)
    
    inv_number_re_pattern = r'INVOICE #(\D+)'
    inv_date_re_pattern = r'DATE: (\D{2}/\D{2}/\D{4})'

    match_number = re.search(inv_number_re_pattern, pdf_text)
    match_date = re.search(inv_date_re_pattern, pdf_text)

    
    if match_number:
        invoice_number = match_number.group(1)
        ws['A{}'.format(last_empty_line)] = invoice_number
    else:
        ws['A{}'.format(last_empty_line)] = "Invoice number not match."

    if match_date:
        invoice_date = match_date.group(1)       
    else:
        ws['A{}'.format(last_empty_line)] = "Invoice date not match."
    
    ws['C{}'.format(last_empty_line)] = file
    ws['D{}'.format(last_empty_line)] = "Completed!"

    last_empty_line += 1

full_now = str(datetime.now()).replace(':','-')
dot_index = full_now.index('.')
now = full_now[:dot_index]
    
wb.save('Invoices - {}.xlsx'.format(now))

print(f'Fim do script - {now}')